import json
import os
from dataclasses import dataclass
from os.path import join
from typing import Callable, Optional, Union

import openpyxl
from openpyxl.cell import Cell
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from tqdm import tqdm

from src.config import BRANCH_MAPPINGS, JSON_FOLDER, REPORTS_FOLDER

Headers = list[str]
Rows = list[dict[str, Optional[str]]]
ExcelRow = tuple[Cell]


@dataclass
class Report:
    report_name: str
    headers_idx: int = 0
    data_idx = 0
    subheaders_exist: bool = False
    parse_rows: Callable[[Worksheet, int, Headers], Rows] = None

    def __post_init__(self):
        if self.report_name == 'Z_160_SHTAT_RASTANOVKA_V_00':
            self.headers_idx, self.data_idx = 3, 6
            self.subheaders_exist = True
            self.parse_rows = parse_z_160_shtat_rastanovka_v_rows
        elif self.report_name == 'Z_160_DISEMPLOYEE_00':
            self.headers_idx, self.data_idx = 4, 5
            self.parse_rows = parse_z_160_dismeployee_rows
        elif self.report_name == 'Z_160_HREMPLOYTAKETOWORK_00':
            self.headers_idx, self.data_idx = 5, 7
            self.parse_rows = parse_z_160_hremploytaketowork_rows
        elif 'Z_160_PR_FORMOBWVEDZP' in self.report_name:
            self.headers_idx, self.data_idx = 3, 5
            self.parse_rows = parse_z_160_pr_formobwvedzp_rows
        else:
            raise ValueError(f'Unknown report name: {self.report_name}')


def get_branch(branch_name: str) -> str:
    branch_name = branch_name.strip().lower()

    branch_mapping: dict[str, Union[str, list[str]]]
    branch = next((branch_mapping['branch'] for branch_mapping in BRANCH_MAPPINGS
                   if branch_name in branch_mapping['long_alias'].lower()), None)
    if not branch:
        raise ValueError(f'Unknown branch name: {branch_name}')
    return branch


def is_row_empty(row: ExcelRow) -> bool:
    return all(not (bool(value.strip()) if isinstance(value := cell.value, str) else bool(value)) for cell in row)


def parse_headers(sheet: Worksheet, headers_idx: int, subheaders_exist: bool = False) -> Headers:
    sheet_headers = []
    for col in sheet.iter_cols(min_row=headers_idx, max_row=headers_idx, min_col=1):
        for header_cell in col:
            cell_value = header_cell.value
            if not cell_value:
                break
            if subheaders_exist:
                if header_cell.coordinate == 'L3':
                    for column in range(5):
                        cell_beneath = sheet.cell(row=header_cell.row + 1,
                                                  column=header_cell.column + column)
                        sheet_headers.append(f'{cell_value}_{cell_beneath.value}')
                    continue
            sheet_headers.append(header_cell.value)
    return sheet_headers


def parse_z_160_shtat_rastanovka_v_rows(sheet: Worksheet, data_idx: int, headers: Headers) -> Rows:
    rows = []
    merged_cells = sheet.merged_cells
    current_branch = ''

    row: ExcelRow
    for i, row in enumerate(sheet.iter_rows(min_row=data_idx)):
        if is_row_empty(row) or not (row[0].value or row[1].value):
            continue

        if merged_cells and row[0].coordinate in merged_cells:
            value = row[0].value
            if not value.startswith(' '):
                current_branch = value
            continue
        else:
            value = row[0].value
            if row[1].value is None and value is not None:
                if not value.startswith(' '):
                    current_branch = value
                continue
        staff_data_row = {'Филиал': get_branch(current_branch)}
        for header, cell in zip(headers, row):
            staff_data_row[header] = cell.value
        rows.append(staff_data_row)

    return rows


def parse_z_160_dismeployee_rows(sheet: Worksheet, data_idx: int, headers: Headers) -> Rows:
    rows = []
    merged_cells = sheet.merged_cells
    current_branch = ''

    row: ExcelRow
    for i, row in enumerate(sheet.iter_rows(min_row=data_idx)):
        if is_row_empty(row):
            continue
        if row[0].value is not None and 'Итого' in row[0].value:
            continue

        if merged_cells and row[0].coordinate in merged_cells:
            value = row[0].value
            if not value.startswith(' '):
                current_branch = value
            continue
        staff_data_row = {'Филиал': get_branch(current_branch)}
        for header, cell in zip(headers, row):
            staff_data_row[header] = cell.value
        rows.append(staff_data_row)
    return rows


def parse_z_160_hremploytaketowork_rows(sheet: Worksheet, data_idx: int, headers: Headers) -> Rows:
    rows = []
    current_branch = ''

    row: ExcelRow
    for i, row in enumerate(sheet.iter_rows(min_row=data_idx, min_col=3)):
        if is_row_empty(row):
            continue
        if row[0].value is not None and 'Итого' in row[0].value:
            continue
        value = row[1].value
        if row[0].value is None and value is not None:
            if not value.startswith(' '):
                current_branch = value
            continue
        staff_data_row = {'Филиал': get_branch(current_branch)}
        for header, cell in zip(headers, row):
            staff_data_row[header] = cell.value
        rows.append(staff_data_row)

    return rows


def parse_z_160_pr_formobwvedzp_rows(sheet: Worksheet, data_idx: int, headers: Headers) -> Rows:
    rows = []
    brach_cell = sheet.cell(row=data_idx, column=1)
    current_branch = get_branch(brach_cell.value)

    row: ExcelRow
    for i, row in enumerate(sheet.iter_rows(min_row=data_idx + 1)):
        if is_row_empty(row):
            break

        staff_data_row = {'Филиал': current_branch}
        for header, cell in zip(headers, row):
            staff_data_row[header] = cell.value
        rows.append(staff_data_row)
    return rows


def save_to_json(report_path: str, data: Rows) -> None:
    with open(report_path, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=4)


def parse_report(report: Report) -> None:
    file_path = join(REPORTS_FOLDER, rf'{report.report_name}.xlsx')

    if not os.path.exists(file_path):
        raise FileNotFoundError(f'File {file_path} does not exist')

    workbook: Workbook = openpyxl.load_workbook(file_path)
    sheet: Worksheet = workbook.active

    headers = parse_headers(sheet=sheet, headers_idx=report.headers_idx, subheaders_exist=report.subheaders_exist)
    data = report.parse_rows(sheet=sheet, data_idx=report.data_idx, headers=headers)
    save_to_json(report_path=join(JSON_FOLDER, f'{report.report_name}.json'), data=data)


def task_1t():
    # json_files = os.listdir(JSON_FOLDER)
    # file_name = 'Z_160_PR_FORMOBWVEDZP_18.json'
    # file_path = join(JSON_FOLDER, file_name)
    #
    # with open(file_path, encoding='utf-8') as f:
    #     data = json.load(f)
    #
    # branch_employee_count = len(data)
    # factual_branch_employee_count = branch_employee_count
    #
    # for row in data:
    #     if row['Состояние'] in ['Уволен', 'Уволен, Отпуск по уходу за ребенком (не достигшего 3-х лет)',
    #                             'В отпуске, Отпуск по уходу за ребенком (не достигшего 3-х лет)']:
    #         factual_branch_employee_count -= 1
    #
    # print(branch_employee_count, factual_branch_employee_count)




    file_path = r"C:\Users\robot.ad\Desktop\robots\1t_otbasy\json\4 квартал 2023\Z_160_PR_FORMOBWVEDZP_18.json"

    with open(file_path, encoding='utf-8') as f:
        data = json.load(f)

    branch_employee_count = 0
    total_salary_fund = 0
    total_work_hours = 0
    employees = []
    for record in data:
        status = record['Состояние']
        if status in ['Уволен',
                      'В отпуске, Отпуск по уходу за ребенком (не достигшего 3-х лет)',
                      'Уволен, Отпуск по уходу за ребенком (не достигшего 3-х лет)'
                      # 'Работающий, Отпуск по уходу за ребенком (не достигшего 3-х лет)',
                      ]:
            continue
        else:
            total_salary = record.get('Выплачено доходов')
            if isinstance(total_salary, str):
                # Замена запятой на точку для корректного преобразования в число
                total_salary = float(total_salary.replace(',', '.'))
            elif isinstance(total_salary, int):
                total_salary = float(total_salary)
            total_salary_fund += total_salary

            total_work_hours += int(record.get('Кол-во часов', 0)) if record.get('Кол-во часов', 0) else 0
        employees.append(record['Сотрудник'])
        # else:
        #     if not row['Кол-во часов']:
        #         continue
        branch_employee_count += 1
    total_salary_fund_thousand_tenge = total_salary_fund / 1000
    print(branch_employee_count, total_salary_fund_thousand_tenge, total_work_hours)
    #
    #
    #
    #
    #
    # file_name = 'Z_160_SHTAT_RASTANOVKA_V_00.json'
    # file_path = join(JSON_FOLDER, file_name)
    #
    # with open(file_path, encoding='utf-8') as f:
    #     data = json.load(f)
    #
    # successful_hits = 0
    # women_count = 0
    # for record in data:
    #     branch, full_name, status, emp_type, sex = record['Филиал'], record['ФИО'], record['Статус работника'], record['Тип работника'], record['Пол']
    #     if branch == '18':
    #         if full_name in employees:
    #             successful_hits += 1
    #             women_count += 1 if sex == 'Женщина' else 0
    #         # branch_employee_count += 1
    # print(successful_hits, len(employees), women_count)
    # pass


    # file_path = r"C:\Users\robot.ad\Desktop\robots\1t_otbasy\json\4 квартал 2023\Z_160_HREMPLOYTAKETOWORK_00.json"
    #
    # with open(file_path, encoding='utf-8') as f:
    #     data = json.load(f)
    #
    # hired_emp_count = 0
    # for record in data:
    #     branch = record['Филиал']
    #     if branch == '18':
    #         hired_emp_count += 1
    # print(hired_emp_count)

    # file_path = r"C:\Users\robot.ad\Desktop\robots\1t_otbasy\json\4 квартал 2023\Z_160_DISEMPLOYEE_00.json"
    #
    # with open(file_path, encoding='utf-8') as f:
    #     data = json.load(f)
    #
    # voluntary_firing_count = 0
    # other_firing_count = 0
    # unique_reasons_for_firing = set()
    # for record in data:
    #     branch = record['Филиал']
    #     if branch == '18':
    #         reason = record['Статья']
    #         if (reason == 'п. 5 ст. 49 Трудового Кодекса Республики Казахстан'
    #                 or reason == 'п. 5 ст. 49 Трудового Кодекса Республики Казахстан; (расторжение трудового договора по инициативе работника)'):
    #             voluntary_firing_count += 1
    #         elif reason == 'п. 1 ст. 49 Трудового Кодекса Республики Казахстан. (Расторжение трудового договора по соглашению сторон)':
    #             other_firing_count += 1
    #         unique_reasons_for_firing.add(record['Статья'])
    # print(unique_reasons_for_firing)
    # print(voluntary_firing_count, other_firing_count)


def main():
    # os.makedirs(JSON_FOLDER, exist_ok=True)
    # reports = os.listdir(REPORTS_FOLDER)
    # for file_name in tqdm(reports, smoothing=0, desc='Parsing reports'):
    #     report_name = file_name.split('.')[0]
    #     parse_report(report=Report(report_name=report_name))

    task_1t()


if __name__ == '__main__':
    main()
